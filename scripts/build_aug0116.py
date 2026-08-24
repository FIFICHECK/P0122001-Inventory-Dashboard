#!/usr/bin/env python3
"""Build the 2026-08-01..16 report from jerry-dashboard data (derived).

jerry-dashboard has August 2026 as a full-month figure only (GMV $807,511.10).
Daily reports 08-17..23 exist (user uploads), sum $142,806.60.
So 08-01..16 = $807,511.10 - $142,806.60 = $664,704.50 (derived, NOT actual daily data).

This builds reports/order_reports/P0122001_GMV_Period_20260801-20260816.xlsx
using jerry's August SKU-level data scaled proportionally so the total matches
$664,704.50, and registers it in data/gmv_monthly_totals.json under '20260801-16'.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

AUG_TOTAL = 807511.10
DAILY_17_23 = 142806.60
PERIOD_TOTAL = round(AUG_TOTAL - DAILY_17_23, 2)   # 664704.50
SCALE = PERIOD_TOTAL / AUG_TOTAL                    # ~0.8231

# ---------- 1. Load jerry August SKU data ----------
src = os.path.expanduser('~/Downloads/sku_data_full.json')
if not os.path.exists(src):
    import urllib.request
    with urllib.request.urlopen('https://fificheck.github.io/jerry-dashboard/sku_data_full.json') as resp:
        jerry = json.load(resp)
else:
    jerry = json.load(open(src, encoding='utf-8'))

aug = {}
for e in jerry:
    sc = e.get('sc', '')
    if sc.startswith('P0122001_S_') and e.get('m') == '2026-08':
        aug[sc] = {'name': e.get('sn', ''), 'gmv': float(e.get('gmv') or 0), 'qty': int(e.get('qty') or 0)}
print(f"jerry 2026-08: {len(aug)} SKUs | scale factor {SCALE:.4f}")

# ---------- 2. Write period xlsx (SKU rows scaled to 08-01..16 total) ----------
out = 'reports/order_reports/P0122001_GMV_Period_20260801-20260816.xlsx'
headers = ['Virtual Store / Store','SKU / Bundle','brand_chi','SKU / Bundle Name',
           'Primary Sub Cat 1 Code','Primary Sub Cat 2 Code','Primary Sub Cat 3 Code',
           'Primary Sub Cat 4 Code','p_rmcode','GMV','Cust #','Parent Order #','quantity']
wb = Workbook()
sh = wb.active
sh.title = 'By SKU'
sh.cell(row=1, column=1, value='P0122001 (SKECHERS) — 2026-08-01..16 GMV (derived: Aug full $807,511.10 − 08-17..23 daily $142,806.60)').font = Font(bold=True, size=11)
for c, htxt in enumerate(headers, start=1):
    sh.cell(row=2, column=c, value=htxt).font = Font(bold=True)

r_out = 3
gmv_tot = 0.0
qty_tot = 0
for sku, d in sorted(aug.items(), key=lambda x: -x[1]['gmv']):
    scaled_gmv = round(d['gmv'] * SCALE, 2)
    scaled_qty = round(d['qty'] * SCALE)
    sh.cell(row=r_out, column=2, value=sku)
    sh.cell(row=r_out, column=4, value=d['name'])
    sh.cell(row=r_out, column=10, value=scaled_gmv)
    sh.cell(row=r_out, column=13, value=scaled_qty)
    gmv_tot += scaled_gmv
    qty_tot += scaled_qty
    r_out += 1
sh.cell(row=r_out, column=4, value='合計').font = Font(bold=True)
sh.cell(row=r_out, column=10, value=round(gmv_tot, 2)).font = Font(bold=True)
sh.cell(row=r_out, column=13, value=qty_tot).font = Font(bold=True)
wb.save(out)
print(f"written {out}: {r_out-3} SKUs | GMV ${gmv_tot:,.2f} (target ${PERIOD_TOTAL:,.2f}) | qty {qty_tot}")

# ---------- 3. Register in sidecar ----------
sidecar = 'data/gmv_monthly_totals.json'
totals = json.load(open(sidecar, encoding='utf-8')) if os.path.exists(sidecar) else {}
totals['20260801-16'] = round(gmv_tot, 2)
json.dump(totals, open(sidecar, 'w'), ensure_ascii=False, indent=1)
print("sidecar:", totals)
