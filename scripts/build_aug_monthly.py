#!/usr/bin/env python3
"""Build the 2026-08 monthly GMV report from jerry-dashboard sku_data_full.json.

jerry-dashboard has no daily data for 2026-08-01..16 — only the full month
(1,430 P0122001 SKU records, GMV $807,511.10). This generates
reports/order_reports/P0122001_GMV_Monthly_202608.xlsx in the same format as the
split_gmv_monthly.py outputs (SKU detail + GMV/qty totals) + updates data/gmv_monthly_totals.json
so the Daily Order Report download table gets a 2026-08月全月 row.
"""
import json, os, glob
from openpyxl import Workbook
from openpyxl.styles import Font

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

# ---------- 1. Load jerry August data ----------
src = os.path.expanduser('~/Downloads/sku_data_full.json')
if not os.path.exists(src):
    import urllib.request
    with urllib.request.urlopen('https://fificheck.github.io/jerry-dashboard/sku_data_full.json') as resp:
        jerry = json.load(resp)
else:
    jerry = json.load(open(src, encoding='utf-8'))

aug = {}
for e in jerry:
    sc = e.get('sc', '')
    if sc.startswith('P0122001_S_') and e.get('m') == '2026-08':
        aug[sc] = {
            'name': e.get('sn', ''),
            'gmv': round(float(e.get('gmv') or 0), 2),
            'qty': int(e.get('qty') or 0),
        }
print(f"jerry 2026-08: {len(aug)} SKUs")

# ---------- 2. Write monthly xlsx (same format as split_gmv_monthly) ----------
out = 'reports/order_reports/P0122001_GMV_Monthly_202608.xlsx'
headers = ['Virtual Store / Store','SKU / Bundle','brand_chi','SKU / Bundle Name',
           'Primary Sub Cat 1 Code','Primary Sub Cat 2 Code','Primary Sub Cat 3 Code',
           'Primary Sub Cat 4 Code','p_rmcode','GMV','Cust #','Parent Order #','quantity']
wb = Workbook()
sh = wb.active
sh.title = 'By SKU'
sh.cell(row=1, column=1, value='P0122001 (SKECHERS) — 2026-08 GMV Report (jerry-dashboard source)').font = Font(bold=True, size=12)
for c, htxt in enumerate(headers, start=1):
    sh.cell(row=2, column=c, value=htxt).font = Font(bold=True)

r_out = 3
gmv_tot = 0.0
qty_tot = 0
for sku, d in sorted(aug.items(), key=lambda x: -x[1]['gmv']):
    sh.cell(row=r_out, column=2, value=sku)
    sh.cell(row=r_out, column=4, value=d['name'])
    sh.cell(row=r_out, column=10, value=d['gmv'])
    sh.cell(row=r_out, column=13, value=d['qty'])
    gmv_tot += d['gmv']
    qty_tot += d['qty']
    r_out += 1
sh.cell(row=r_out, column=4, value='合計').font = Font(bold=True)
sh.cell(row=r_out, column=10, value=round(gmv_tot, 2)).font = Font(bold=True)
sh.cell(row=r_out, column=13, value=qty_tot).font = Font(bold=True)
wb.save(out)
print(f"written {out}: {r_out-3} SKUs | GMV ${gmv_tot:,.2f} | qty {qty_tot}")

# ---------- 3. Update sidecar ----------
sidecar = 'data/gmv_monthly_totals.json'
totals = json.load(open(sidecar, encoding='utf-8')) if os.path.exists(sidecar) else {}
totals['202608'] = round(gmv_tot, 2)
json.dump(totals, open(sidecar, 'w'), ensure_ascii=False, indent=1)
print("updated data/gmv_monthly_totals.json:", totals)
