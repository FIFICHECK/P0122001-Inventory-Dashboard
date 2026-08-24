#!/usr/bin/env python3
"""Build data/report_manifest.json for the Report tab download tables.

Scans reports/ for inventory CSVs and reports/order_reports/ for order XLSX files,
computes date/time/SKU-count/GMV where derivable, and emits:
{
  "inventory_reports": [{"date": "2026-08-21", "time": "1200", "skus": 40436, "file": "inventory_report_20260821_1200.csv", "size": 12252314}],
  "order_reports": [{"date": "2026-08-23", "time": "235959", "gmv": 12345.67, "file": "ECOM-...xlsx", "size": 87184}],
  "updated_at": "..."
}
"""
import os, re, glob, json, csv, datetime

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

def parse_inventory_file(name):
    m = re.match(r'inventory_report_(\d{8})_(\d{4})\.csv$', name)
    if not m:
        return None
    date_raw, time_raw = m.group(1), m.group(2)
    date = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]}"
    # count SKU rows (strip 5 header lines)
    path = os.path.join('reports', name)
    try:
        lines = open(path, encoding='utf-8-sig', errors='replace').read().split('\n')
        # data rows: after header line (index 5). Count non-empty
        skus = sum(1 for l in lines[6:] if l.strip())
    except Exception:
        skus = 0
    return {'date': date, 'time': time_raw, 'skus': skus, 'file': name, 'size': os.path.getsize(path)}

def read_daily_gmv(name):
    """Read GMV from an ECOM daily order report xlsx (header F2 or sum col AA).
    Uses openpyxl via python3.12 site-packages if available; returns None otherwise."""
    try:
        import openpyxl
    except ImportError:
        import sys as _sys
        _sys.path.insert(0, '/home/snkwok/.local/lib/python3.12/site-packages')
        try:
            import openpyxl
        except ImportError:
            return None
    try:
        wb = openpyxl.load_workbook(os.path.join('reports/order_reports', name), data_only=True)
        ws = wb.active
        if ws is None:
            return None
        hdr = ws.cell(row=2, column=6).value
        if isinstance(hdr, (int, float)):
            return round(float(hdr), 2)
        total = sum(ws.cell(row=r, column=27).value for r in range(5, ws.max_row + 1)
                    if isinstance(ws.cell(row=r, column=27).value, (int, float)))
        return round(total, 2) if total else None
    except Exception:
        return None

def parse_order_file(name):
    m = re.match(r'ECOM-MMSNG_DAILY_ORDER_P0122001_(\d{8})(\d{6})\.xlsx$', name)
    if not m:
        m = re.match(r'ECOM-MMSNG_DAILY_ORDER_P0122001_(\d{8})(\d{6})\.xls$', name)
    if m:
        date_raw, time_raw = m.group(1), m.group(2)
        date = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]}"
        path = os.path.join('reports/order_reports', name)
        return {'date': date, 'time': time_raw, 'gmv': read_daily_gmv(name),
                'file': name, 'size': os.path.getsize(path)}
    # Monthly GMV reports (split from GP Report Excel): P0122001_GMV_Monthly_YYYYMM.xlsx
    m2 = re.match(r'P0122001_GMV_Monthly_(\d{6})\.xlsx$', name)
    if m2:
        y, mo = m2.group(1)[:4], m2.group(1)[4:6]
        date = f"{y}-{mo}-01"
        # GMV total from sidecar JSON (written by split_gmv_monthly.py; system python3 lacks openpyxl)
        gmv = None
        try:
            gmv_map = json.load(open('data/gmv_monthly_totals.json', encoding='utf-8'))
            gmv = gmv_map.get(m2.group(1))
        except Exception:
            gmv = None
        path = os.path.join('reports/order_reports', name)
        return {'date': date, 'date_label': f"{y}-{mo}月全月", 'time': '235959',
                'gmv': round(float(gmv), 2) if isinstance(gmv, (int, float)) else None,
                'file': name, 'size': os.path.getsize(path)}
    # Period reports (derived from jerry-dashboard): P0122001_GMV_Period_YYYYMMDD-YYYYMMDD.xlsx
    m3 = re.match(r'P0122001_GMV_Period_(\d{8})-(\d{8})\.xlsx$', name)
    if m3:
        d1, d2 = m3.group(1), m3.group(2)
        date = f"{d2[:4]}-{d2[4:6]}-{d2[6:8]}"
        label = f"{d1[:4]}-{d1[4:6]}-{d1[6:8]} ~ {d2[4:6]}-{d2[6:8]}"
        gmv = None
        try:
            gmv_map = json.load(open('data/gmv_monthly_totals.json', encoding='utf-8'))
            gmv = gmv_map.get(f"{d1}-{d2[6:8]}")   # key format: YYYYMMDD-DD (e.g. 20260801-16)
        except Exception:
            gmv = None
        path = os.path.join('reports/order_reports', name)
        return {'date': date, 'date_label': label, 'time': '235959',
                'gmv': round(float(gmv), 2) if isinstance(gmv, (int, float)) else None,
                'file': name, 'size': os.path.getsize(path)}
    return None

inventory = []
for f in sorted(glob.glob('reports/inventory_report_*.csv')):
    e = parse_inventory_file(os.path.basename(f))
    if e:
        inventory.append(e)
inventory.sort(key=lambda x: x['date'], reverse=True)

order = []
for f in sorted(glob.glob('reports/order_reports/*.xlsx') + glob.glob('reports/order_reports/*.xls')):
    e = parse_order_file(os.path.basename(f))
    if e:
        order.append(e)
order.sort(key=lambda x: (x['date'], x['time']), reverse=True)

out = {
    'inventory_reports': inventory,
    'order_reports': order,
    'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
}
json.dump(out, open('data/report_manifest.json', 'w'), ensure_ascii=False, indent=1)
print("inventory reports:", len(inventory), "| order reports:", len(order))
for e in inventory[:3]:
    print("  inv:", e['date'], e['time'], e['skus'], 'SKUs')
for e in order[:10]:
    print("  ord:", e['date'], e['time'], e['gmv'], e['size'], 'bytes', e['file'][:60])
