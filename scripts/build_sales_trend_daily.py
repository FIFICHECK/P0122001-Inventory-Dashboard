#!/usr/bin/env python3
"""Merge Daily Order Report data into salesTrendData daily charts.

Reads all ECOM-MMSNG_DAILY_ORDER_P0122001_*.xlsx in reports/order_reports/
(08-17..23), aggregates per day/hour/day-of-week/SKU/brand, and merges into
data/sales_trend_data.js (which currently has monthly data only from the GP Excel).
Keeps monthly keys untouched; fills the *_daily keys.

Usage: run with python3.12 + openpyxl, then inject via inject_sales.py.
"""
import openpyxl, json, os, glob, re, datetime
from collections import defaultdict

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

# ---------- 1. Read all daily reports ----------
files = sorted(glob.glob('reports/order_reports/ECOM-MMSNG_DAILY_ORDER_P0122001_*.xlsx'))
print(f"daily reports: {len(files)}")
for f in files:
    print("  ", os.path.basename(f))

gmv_by_date = defaultdict(float)
qty_by_date = defaultdict(int)
orders_by_date = defaultdict(int)          # parent order count per date
gmv_by_hour = defaultdict(float)
gmv_by_dow = defaultdict(float)
gmv_by_sku = defaultdict(lambda: defaultdict(float))   # date -> sku -> gmv
qty_by_sku = defaultdict(lambda: defaultdict(int))     # date -> sku -> qty
sku_names = {}
gmv_by_brand = defaultdict(lambda: defaultdict(float))

DATE_KEYS = []
for f in files:
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    date = None
    parent_orders = set()
    for r in range(6, ws.max_row + 1):
        od = ws.cell(row=r, column=7).value
        if od is None:
            continue
        date = str(od)[:10]
        if date not in DATE_KEYS:
            DATE_KEYS.append(date)
        # hour: col 8 like '00:40:37'
        tm = str(ws.cell(row=r, column=8).value or '')
        hour = int(tm[:2]) if len(tm) >= 2 and tm[:2].isdigit() else 0
        sku = str(ws.cell(row=r, column=18).value or '').strip()
        name = str(ws.cell(row=r, column=22).value or ws.cell(row=r, column=21).value or '').strip()
        qty = ws.cell(row=r, column=24).value or 0
        total = ws.cell(row=r, column=27).value or 0
        oid = str(ws.cell(row=r, column=5).value or '')   # Sub-Order ID
        parent_orders.add(oid)
        gmv_by_date[date] += float(total)
        qty_by_date[date] += int(qty or 0)
        gmv_by_hour[hour] += float(total)
        dow = datetime.date.fromisoformat(date).weekday()
        gmv_by_dow[dow] += float(total)
        if sku:
            gmv_by_sku[date][sku] += float(total)
            qty_by_sku[date][sku] += int(qty or 0)
            sku_names[sku] = name or sku
        gmv_by_brand[date]['SKECHERS'] += float(total)
    orders_by_date[date] = len(parent_orders)

DATE_KEYS.sort()
print(f"\ndates: {DATE_KEYS}")
print(f"daily GMV: { {d: round(gmv_by_date[d],2) for d in DATE_KEYS} }")

# ---------- 2. Load existing sales_trend_data.js ----------
js = open('data/sales_trend_data.js', encoding='utf-8').read()
m = re.search(r'const salesTrendData = (\{.*\});', js, re.S)
if not m:
    raise SystemExit('!! cannot find salesTrendData in data/sales_trend_data.js')
sd = json.loads(m.group(1))

# ---------- 3. Merge daily data ----------
# gmv_by_date: keep existing (empty) + fill
sd['gmv_by_date'] = {'labels': DATE_KEYS, 'data': [round(gmv_by_date[d], 2) for d in DATE_KEYS]}
sd['orders_by_date'] = {'labels': DATE_KEYS, 'data': [orders_by_date[d] for d in DATE_KEYS]}
sd['gmv_by_hour'] = {'labels': [], 'hours': [f"{h:02d}" for h in range(24)], 'data': {}}
sd['gmv_by_hour']['data'] = {f"{h:02d}": round(gmv_by_hour.get(h, 0), 2) for h in range(24)}
# gmv_by_day_of_week: labels like "Monday (2026-08-17)" — JS filters by date substring
dow_names = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
dow_labels = [f"{dow_names[datetime.date.fromisoformat(d).weekday()]} ({d})" for d in DATE_KEYS]
dow_data = [round(gmv_by_date[d], 2) for d in DATE_KEYS]
sd['gmv_by_day_of_week'] = {'labels': dow_labels, 'data': dow_data}
sd['gmv_by_date_hour'] = {'labels': DATE_KEYS, 'hours': [f"{h:02d}" for h in range(24)], 'data': {}}

# SKU daily: top SKUs by total GMV across the window
sku_totals = defaultdict(float)
for d, m2 in gmv_by_sku.items():
    for s, v in m2.items():
        sku_totals[s] += v
# All SKUs (no display limit — user preference)
top_skus = [s for s, _ in sorted(sku_totals.items(), key=lambda x: -x[1])]
sd['gmv_by_sku_daily'] = {'labels': DATE_KEYS, 'skus': top_skus,
                          'data': [[round(gmv_by_sku[d].get(s, 0), 2) for s in top_skus] for d in DATE_KEYS]}
sd['qty_by_sku_daily'] = {'labels': DATE_KEYS, 'skus': top_skus,
                          'data': [[qty_by_sku[d].get(s, 0) for s in top_skus] for d in DATE_KEYS]}
sd['sku_name_map'] = {**sd.get('sku_name_map', {}), **{s: sku_names.get(s, s) for s in top_skus}}
sd['sku_product_name_map'] = sd['sku_name_map']
sd['sku_brand_map'] = {**sd.get('sku_brand_map', {}), **{s: 'SKECHERS' for s in top_skus}}

sd['gmv_by_brand_daily'] = {'labels': DATE_KEYS, 'brands': ['SKECHERS'],
                            'data': [[round(gmv_by_brand[d].get('SKECHERS', 0), 2)] for d in DATE_KEYS]}

# Update summary: this_month/last_month with actual daily data when available
summ = sd['summary']
if DATE_KEYS:
    first, last = DATE_KEYS[0], DATE_KEYS[-1]
    summ['daily_range'] = f"{first} ~ {last}"
    summ['daily_gmv_total'] = round(sum(gmv_by_date.values()), 2)
    summ['daily_orders_total'] = int(sum(orders_by_date.values()))
    summ['daily_qty_total'] = int(sum(qty_by_date.values()))
    summ['avg_order_value_daily'] = round(summ['daily_gmv_total'] / summ['daily_orders_total'], 2) if summ['daily_orders_total'] else 0
    # this_month: GP Excel has no Aug orders — fill from daily reports (7 days available)
    if summ.get('this_month') and summ['this_month'].get('orders', 0) == 0:
        summ['this_month']['orders'] = summ['daily_orders_total']
        summ['this_month']['avg'] = summ['avg_order_value_daily']

# available_months: add the daily dates' month if missing (2026-08 already there)
for d in DATE_KEYS:
    mth = d[:7]
    if mth not in sd.get('available_months', []):
        sd['available_months'].append(mth)

# ---------- 4. Write back ----------
out = js[:m.start()] + json.dumps(sd, ensure_ascii=False, separators=(',', ':')) + js[m.end():]
# keep the const declaration line intact (it's inside m.group(1) boundary handling)
# simplest: rebuild full file
hdr = "// Auto-generated sales trend data (GP Report Excel + daily order reports)\n"
hdr += "// Generated: " + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "\n\n"
js_new = hdr + "const salesTrendData = " + json.dumps(sd, ensure_ascii=False, separators=(',', ':')) + ";\n"
open('data/sales_trend_data.js', 'w', encoding='utf-8').write(js_new)
print(f"\nwritten data/sales_trend_data.js ({len(js_new)//1024} KB)")
print(f"gmv_by_date: {sd['gmv_by_date']['data']}")
print(f"orders_by_date: {sd['orders_by_date']['data']}")
print(f"hour data: {sd['gmv_by_hour']['data']}")
print(f"dow: {sd['gmv_by_day_of_week']['data']}")
print(f"sku daily top5: {sd['gmv_by_sku_daily']['skus'][:5]}")
print(f"brand daily: {sd['gmv_by_brand_daily']['data']}")
