#!/usr/bin/env python3
"""Build P0122001 sales_trend_data.js from the GP Report Excel (By SKU By Month).

Source: reports/P0122001_GMV_By_SKU_By_Month_202601-202607.xlsx (Tableau GP Report export,
uploaded by user 2026-08-24). Columns: GMV (10-16), Cust# (17-23), Parent Order# (24-30),
quantity (31-37) for 2026-01..07. 4,945 SKU rows — full SKU coverage (jerry-dashboard only had Top 50).

Output: data/sales_trend_data.js with the B0961005 inline format:
  gmv_by_month, qty_by_month, orders_by_month (NEW), customers_by_month (NEW),
  gmv_by_sku_monthly (all SKUs × 8 months, Top 1000 for payload), qty_by_sku_monthly,
  gmv_by_brand_monthly, sku_name_map, sku_brand_map, available_months, summary
NOTE: 2026-08 (August) still comes from jerry-dashboard monthly data (sku_data_full.json) since
the Excel covers only Jan-Jul. Inject via inject_sales.py.
"""
import openpyxl, json, os, re, glob
import datetime as _dt

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

MONTHS = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06","2026-07"]

# ---------- 1. Read Excel ----------
xlsx = glob.glob('reports/P0122001_GMV_By_SKU_By_Month_*.xlsx')
if not xlsx:
    raise SystemExit("!! no GMV Excel found in reports/")
wb = openpyxl.load_workbook(xlsx[0], data_only=True)
ws = wb["By SKU By Month"]

sku_gmv = {}   # sku -> [7 gmv values]
sku_qty = {}   # sku -> [7 qty values]
sku_names = {} # sku -> name
def _num(v):
    """Cell value -> float, None-safe."""
    return float(v) if isinstance(v, (int, float)) else 0.0

def _inum(v):
    """Cell value -> int, None-safe."""
    return int(v) if isinstance(v, (int, float)) else 0

for r in range(3, ws.max_row + 1):
    sku = ws.cell(row=r, column=2).value
    if not sku:
        continue
    gmv = [_num(ws.cell(row=r, column=c).value) for c in range(10, 17)]
    qty = [_inum(ws.cell(row=r, column=c).value) for c in range(31, 38)]
    if any(gmv) or any(qty):
        sku_gmv[sku] = [round(x, 2) for x in gmv]
        sku_qty[sku] = qty
        sku_names[sku] = (ws.cell(row=r, column=4).value or '')

monthly_gmv = {m: 0.0 for m in MONTHS}
monthly_qty = {m: 0 for m in MONTHS}
orders_by_month = {m: 0 for m in MONTHS}
customers_by_month = {m: 0 for m in MONTHS}
for i, m in enumerate(MONTHS):
    monthly_gmv[m] = round(sum(v[i] for v in sku_gmv.values()), 2)
    monthly_qty[m] = sum(v[i] for v in sku_qty.values())
    orders_by_month[m] = sum(_inum(ws.cell(row=r, column=24 + i).value) for r in range(3, ws.max_row + 1))
    customers_by_month[m] = sum(_inum(ws.cell(row=r, column=17 + i).value) for r in range(3, ws.max_row + 1))

# ---------- 2. August from jerry-dashboard ----------
def load_jerry():
    # Prefer the committed snapshot so Aug GMV matches the Daily Order Report tab
    # ($807,511.10) — live jerry may drift (2026-08-24: live showed $832,752.30).
    p = os.path.expanduser('~/P0122001-Inventory-Dashboard/data/jerry_sku_data_full.json')
    if os.path.exists(p):
        return json.load(open(p, encoding='utf-8'))
    p2 = os.path.expanduser('~/Downloads/sku_data_full.json')
    if os.path.exists(p2):
        return json.load(open(p2, encoding='utf-8'))
    # fallback: fetch live
    import urllib.request
    with urllib.request.urlopen('https://fificheck.github.io/jerry-dashboard/sku_data_full.json') as resp:
        return json.load(resp)

jerry = load_jerry()
aug_gmv = {}
aug_qty = {}
aug_names = {}
for e in jerry:
    sc = e.get('sc', '')
    if sc.startswith('P0122001_S_') and e.get('m') == '2026-08':
        aug_gmv[sc] = aug_gmv.get(sc, 0) + (e.get('gmv') or 0)
        aug_qty[sc] = aug_qty.get(sc, 0) + (e.get('qty') or 0)
        aug_names[sc] = e.get('sn', '')
aug_total_gmv = round(sum(aug_gmv.values()), 2)
aug_total_qty = int(sum(aug_qty.values()))
print(f"jerry Aug: {len(aug_gmv)} SKUs, GMV ${aug_total_gmv:,.2f}, qty {aug_total_qty}")

# ---------- 3. Merge: all months 2026-01..08 ----------
ALL_MONTHS = MONTHS + ["2026-08"]
all_skus = set(sku_gmv.keys()) | set(aug_gmv.keys())
print(f"total unique SKUs (excel + aug): {len(all_skus)}")

merged_gmv = {}
merged_qty = {}
merged_names = {}
for sku in all_skus:
    eg = sku_gmv.get(sku, [0]*7)
    eq = sku_qty.get(sku, [0]*7)
    merged_gmv[sku] = eg + [round(aug_gmv.get(sku, 0), 2)]
    merged_qty[sku] = eq + [int(aug_qty.get(sku, 0))]
    merged_names[sku] = sku_names.get(sku) or aug_names.get(sku) or sku

# monthly totals incl Aug
for m, val in [("2026-08", aug_total_gmv)]:
    monthly_gmv[m] = val
for m, val in [("2026-08", aug_total_qty)]:
    monthly_qty[m] = val
orders_by_month["2026-08"] = None   # no orders data for Aug yet (jerry has none)
customers_by_month["2026-08"] = None

# ---------- 3b. GMV Target & Runrate (user-defined monthly targets, 2026) ----------
TARGET_MONTHS = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06","2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]
GMV_TARGETS = [1200000, 950000, 950000, 950000, 1300000, 900000, 900000, 1000000, 950000, 1300000, 1200000, 1100000]
YEAR_TARGET = 12700000
gmv_target = {
    'labels': TARGET_MONTHS,
    'targets': GMV_TARGETS,
    'actual': [round(monthly_gmv.get(m, 0), 2) for m in TARGET_MONTHS],  # 0 for future months
    'year_target': YEAR_TARGET,
}

# ---------- 4. Build chart data ----------
gmv_by_month = {'labels': ALL_MONTHS, 'data': [round(monthly_gmv[m], 2) for m in ALL_MONTHS]}
qty_by_month = {'labels': ALL_MONTHS, 'data': [monthly_qty[m] for m in ALL_MONTHS]}
orders_by_month_out = {'labels': ALL_MONTHS, 'data': [orders_by_month[m] if orders_by_month[m] is not None else None for m in ALL_MONTHS]}
customers_by_month_out = {'labels': ALL_MONTHS, 'data': [customers_by_month[m] if customers_by_month[m] is not None else None for m in ALL_MONTHS]}

# All SKUs by total GMV (no display limit — user preference)
sku_totals = {sku: sum(v) for sku, v in merged_gmv.items()}
top_skus = sorted(sku_totals, key=lambda x: -sku_totals[x])
sku_monthly_data = [[round(merged_gmv[s][i], 2) for s in top_skus] for i in range(8)]  # Month×SKU
qty_sku_monthly_data = [[merged_qty[s][i] for s in top_skus] for i in range(8)]
gmv_sku_monthly = {'labels': ALL_MONTHS, 'skus': top_skus, 'data': sku_monthly_data}
qty_sku_monthly = {'labels': ALL_MONTHS, 'skus': top_skus, 'data': qty_sku_monthly_data}

gmv_brand_monthly = {
    'labels': ALL_MONTHS,
    'brands': ['SKECHERS'],
    'data': [[round(monthly_gmv[m], 2)] for m in ALL_MONTHS],
}

sku_name_map = {s: merged_names.get(s, s) for s in top_skus}
sku_brand_map = {s: 'SKECHERS' for s in top_skus}

# ---------- 5. Summary ----------
total_gmv = round(sum(monthly_gmv.values()), 2)
total_orders = sum(v for v in orders_by_month.values() if v)
avg_order_value = round(total_gmv / total_orders, 2) if total_orders else 0

def fmt(label, gmv, orders):
    return {
        'label': label,
        'gmv': round(gmv, 2),
        'orders': int(orders) if orders else 0,
        'avg': round(gmv / orders, 2) if orders else 0,
    }

# 動態月份 label（今日 = 當月；上月/上上月跟住 roll）— 修 2026-09-02 發現嘅硬code label bug
_today = _dt.date.today()
def _mlabel(offset_months):
    total = _today.year * 12 + (_today.month - 1) - offset_months
    yy, mm = divmod(total, 12)
    return f"{mm + 1}月 {yy}"

_sorted_months = sorted(ALL_MONTHS)
_latest, _prev, _prev2 = _sorted_months[-1], _sorted_months[-2], _sorted_months[-3]
summary = {
    'this_month': fmt(_mlabel(0), monthly_gmv[_latest], 0),
    'last_month': fmt(_mlabel(1), monthly_gmv[_prev], orders_by_month.get(_prev, 0) or 0),
    'month_before_last': fmt(_mlabel(2), monthly_gmv[_prev2], orders_by_month.get(_prev2, 0) or 0),
    'total_gmv': total_gmv,
    'total_orders': total_orders,
    'avg_order_value': avg_order_value,
    'orders_by_month': orders_by_month_out,
    'customers_by_month': customers_by_month_out,
}

out = {
    'gmv_by_date': {'labels': [], 'data': []},
    'gmv_by_month': gmv_by_month,
    'gmv_by_hour': {'labels': [], 'hours': [f"{h:02d}" for h in range(24)], 'data': {}},
    'gmv_by_day_of_week': {'labels': [], 'data': []},
    'gmv_by_date_hour': {'labels': [], 'hours': [f"{h:02d}" for h in range(24)], 'data': {}},
    'orders_by_date': {'labels': [], 'data': []},
    'gmv_by_sku_daily': {'labels': [], 'skus': [], 'data': []},
    'qty_by_sku_daily': {'labels': [], 'skus': [], 'data': []},
    'gmv_by_sku_monthly': gmv_sku_monthly,
    'qty_by_sku_monthly': qty_sku_monthly,
    'gmv_by_brand_daily': {'labels': [], 'brands': [], 'data': []},
    'gmv_by_brand_monthly': gmv_brand_monthly,
    'sku_name_map': sku_name_map,
    'sku_product_name_map': sku_name_map,
    'sku_brand_map': sku_brand_map,
    'gmv_target': gmv_target,
    'available_months': ALL_MONTHS,
    'summary': summary,
}

js = "// Auto-generated sales trend data (GP Report Excel + jerry-dashboard Aug)\n"
js += "// Generated: " + __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "\n"
js += "// Source: reports/P0122001_GMV_By_SKU_By_Month_202601-202607.xlsx + jerry sku_data_full.json (2026-08)\n\n"
js += "const salesTrendData = " + json.dumps(out, ensure_ascii=False, separators=(',', ':')) + ";\n"

open('data/sales_trend_data.js', 'w', encoding='utf-8').write(js)
print(f"written data/sales_trend_data.js ({len(js)//1024} KB)")
print(f"months: {ALL_MONTHS}")
print(f"gmv_by_month: {gmv_by_month['data']}")
print(f"orders_by_month: {[orders_by_month[m] for m in ALL_MONTHS]}")
print(f"customers_by_month: {[customers_by_month[m] for m in ALL_MONTHS]}")
print(f"top SKUs: {len(top_skus)} | total GMV: ${total_gmv:,.2f} | orders: {total_orders}")
