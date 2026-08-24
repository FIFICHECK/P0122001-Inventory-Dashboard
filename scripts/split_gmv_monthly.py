#!/usr/bin/env python3
"""Split the GP Report Excel (By SKU By Month) into 7 monthly xlsx files.

Each monthly file keeps SKU-level detail for that month:
  reports/order_reports/P0122001_GMV_Monthly_YYYYMM.xlsx
Columns: Store, SKU/Bundle, brand_chi, SKU/Bundle Name, category codes, GMV, Cust#, Parent Order#, quantity
"""
import openpyxl, os, glob, json
from openpyxl.styles import Font

REPO = os.path.expanduser('~/P0122001-Inventory-Dashboard')
os.chdir(REPO)

src = glob.glob('reports/P0122001_GMV_By_SKU_By_Month_*.xlsx')
if not src:
    raise SystemExit('!! no source GMV Excel found')
wb_src = openpyxl.load_workbook(src[0], data_only=True)
ws = wb_src['By SKU By Month']

MONTHS = ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06','2026-07']
# col offsets per month index i: GMV 10+i, Cust 17+i, Order 24+i, qty 31+i
outdir = 'reports/order_reports'
os.makedirs(outdir, exist_ok=True)

headers = ['Virtual Store / Store','SKU / Bundle','brand_chi','SKU / Bundle Name',
           'Primary Sub Cat 1 Code','Primary Sub Cat 2 Code','Primary Sub Cat 3 Code',
           'Primary Sub Cat 4 Code','p_rmcode','GMV','Cust #','Parent Order #','quantity']

created = []
for i, month in enumerate(MONTHS):
    out = os.path.join(outdir, f'P0122001_GMV_Monthly_{month.replace("-","")}.xlsx')
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = 'By SKU'
    # title row
    sh.cell(row=1, column=1, value=f'P0122001 (SKECHERS) — {month} GMV Report (GP Report source)')
    sh.cell(row=1, column=1).font = Font(bold=True, size=12)
    # header row 2
    for c, htxt in enumerate(headers, start=1):
        cell = sh.cell(row=2, column=c, value=htxt)
        cell.font = Font(bold=True)
    # data rows
    r_out = 3
    gmv_tot = 0.0
    cust_tot = 0
    order_tot = 0
    qty_tot = 0
    for r in range(3, ws.max_row + 1):
        sku = ws.cell(row=r, column=2).value
        if not sku:
            continue
        gmv = ws.cell(row=r, column=10 + i).value
        cust = ws.cell(row=r, column=17 + i).value
        order = ws.cell(row=r, column=24 + i).value
        qty = ws.cell(row=r, column=31 + i).value
        if not any(isinstance(v, (int, float)) and v for v in (gmv, cust, order, qty)):
            continue
        vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        vals += [gmv or 0, cust or 0, order or 0, qty or 0]
        for c, v in enumerate(vals, start=1):
            sh.cell(row=r_out, column=c, value=v)
        r_out += 1
        gmv_tot += float(gmv or 0)
        cust_tot += int(cust or 0)
        order_tot += int(order or 0)
        qty_tot += int(qty or 0)
    # totals row
    sh.cell(row=r_out, column=4, value='合計').font = Font(bold=True)
    sh.cell(row=r_out, column=10, value=round(gmv_tot, 2)).font = Font(bold=True)
    sh.cell(row=r_out, column=11, value=cust_tot).font = Font(bold=True)
    sh.cell(row=r_out, column=12, value=order_tot).font = Font(bold=True)
    sh.cell(row=r_out, column=13, value=qty_tot).font = Font(bold=True)
    wb.save(out)
    created.append((month, r_out - 3, round(gmv_tot, 2), order_tot, cust_tot, qty_tot, os.path.getsize(out)))

for month, rows, gmv, orders, custs, qty, size in created:
    print(f"{month}: {rows} SKUs | GMV ${gmv:,.2f} | orders {orders} | cust {custs} | qty {qty} | {size/1024:.0f}KB")

# sidecar JSON for build_report_manifest.py (system python3 lacks openpyxl)
json.dump({m.replace('-', ''): gmv for m, _, gmv, _, _, _, _ in created},
          open('data/gmv_monthly_totals.json', 'w'), ensure_ascii=False, indent=1)
print("written data/gmv_monthly_totals.json")
